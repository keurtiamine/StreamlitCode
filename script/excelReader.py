from openpyxl import load_workbook
# import pandas lib as pd
import pandas as pd



workbook = load_workbook(filename="../files/Bilan Annuel.xlsx")
#get sheet names
print(workbook.sheetnames)

# read by default 1st sheet of an excel file
dataframe1 = pd.read_excel('../files/Bilan Annuel.xlsx',sheet_name = "BA-2022")
print(dataframe1)