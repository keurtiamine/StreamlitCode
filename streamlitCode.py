import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from sympy import symbols, parse_expr
import xlwings as xw
from openpyxl import load_workbook


st.set_page_config(layout="wide")
st.markdown(
    """
    <style>
   
    body {
        background-color: coral; /* Replace #f0f0f0 with your desired background color */
    }
    
    /* Modifier la police et la taille du titre */
    .title {
        font-family: 'Arial', sans-serif;
        font-size: 32px;
        color: #333333;
    }

    /* Modifier la taille des en-têtes de tableau */
    .dataframe th {
        font-size: 16px;
        background-color: #f5f5f5;
    }

    /* Modifier l'arrière-plan du tableau */
    .dataframe {
        background-color: #ffffff;
    }

    /* Ajouter de la marge autour du tableau */
    .dataframe {
        margin: 20px auto;
    }
    </style>
    """,
    unsafe_allow_html=True
)

st.markdown("<h1 style='text-align: center;'>Interface graphique PMP</h1>", unsafe_allow_html=True)


st.title("Affichage intégral de la boucle eau-vapeur")

# Chemin vers l'image
image_path = "SCHEMA BOUCLE EAU VAPEUR.jpg"

# Affichage de l'image
st.image(image_path, caption='Image', use_column_width=True)
calculate_clicked = st.session_state.get('calculate_clicked', False)

# Fonction pour calculer le KPI en fonction de l'expression mathématique et des paramètres sélectionnés
def calculate_kpi(expression, param1, param2, norm_lower, norm_upper):
    # Test pour l'opération de division
    if isinstance(param1, str) or isinstance(param2, str):
        st.write("Les paramètres ne doivent pas être des chaînes de caractère")
        return None
    
    # Évaluation de l'expression mathématique avec les paramètres sélectionnés
    try:
        kpi_result = eval(expression, {'param1': param1, 'param2': param2})
    except Exception as e:
        st.write("Expression mathématique invalide ou erreur de calcul ")
        return None
    
    # Calcul de la différence avec les normes
    diff_lower = np.abs(kpi_result - norm_lower)
    diff_upper = np.abs(kpi_result - norm_upper)
    
    return kpi_result, diff_lower, diff_upper
def calculEtAffichage(param1_values,param2_values):


    # Calcul du KPI et différences avec les normes
    results = [calculate_kpi(expression, param1, param2, norm_lower, norm_upper) for param1, param2 in
               zip(param1_values, param2_values)]

    # Filtrage des résultats valides
    valid_results = [(kpi, diff_lower, diff_upper) for kpi, diff_lower, diff_upper in results if kpi is not None]
    kpi_values, diff_lower_values, diff_upper_values = zip(*valid_results)

    # Affichage du tableau de résultats
    var = st.write('Résultats du KPI :')
    kpi_df = pd.DataFrame({selected_param1: param1_values, selected_param2: param2_values,kpi_name: kpi_values,
                           'Différence inférieure': diff_lower_values, 'Différence supérieure': diff_upper_values})
    st.write(kpi_df)
    excel = kpi_df.to_excel("kpi.xlsx")
    with open('kpi.xlsx', 'rb') as f:
        st.download_button('Télécharger le tableau résultat', f, file_name='kpi.xlsx')

    # Affichage du graphe
    st.write('Graphe du KPI :')
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.plot(np.arange(1, len(kpi_values)+1), kpi_values, marker='o', linestyle='-', linewidth=2, color='blue', label=kpi_name)
    ax.axhline(norm_lower, color='red', linestyle='--', linewidth=2, label='Norme inférieure')
    ax.axhline(norm_upper, color='green', linestyle='--', linewidth=2, label='Norme supérieure')
    ax.set_xlabel('Index')
    ax.set_ylabel('KPI')
    ax.set_title('Évolution du KPI avec les Normes')
    ax.legend()
    ax.grid(True)
    st.pyplot(fig)
    #plot to image
    fig.savefig('plot.png')
    with open('plot.png', 'rb') as f:
        st.download_button('Télécharger le graph', f, file_name='plot.png')

    recalculate_button = st.button("Calculer pour d'autre param")
    if recalculate_button:
        st.session_state['calculate_clicked'] = False
        st.experimental_rerun()

    # Demande à l'utilisateur s'il souhaite effectuer un nouveau calcul
def modify_cell(file_path, cell, value):
    # Open the Excel workbook
    wb = load_workbook(filename=file_path)
    ws = wb.worksheets[0]
    ws[cell] = value
# Interface utilisateur
st.title('Calcul du KPI')

# Chargement des tableaux à partir des fichiers Excel
tableau1_file = st.file_uploader('Charger le tableau 1 (Excel)', type='xlsx')
tableau2_file = st.file_uploader('Charger le tableau 2 (Excel)', type='xlsx')

# Vérification si les fichiers ont été chargés
if tableau1_file is not None and tableau2_file is not None:
    # Lecture des fichiers Excel pour créer les tableaux de données
    tableau1 = pd.read_excel(tableau1_file)
    tableau2 = pd.read_excel(tableau2_file)
    
    # Liste des tableaux disponibles avec leurs noms
    tableaux = {'Tableau 1': tableau1, 'Tableau 2': tableau2}
    
    # Sélection du tableau pour le premier paramètre
    tableau1_choice = st.selectbox('Choisir un tableau pour le premier paramètre', list(tableaux.keys()))
    
    # Récupération du tableau sélectionné pour le premier paramètre
    selected_tableau1 = tableaux[tableau1_choice]
    
    # Affichage du tableau pour le premier paramètre
    st.write('Tableau sélectionné pour le premier paramètre :')
    st.write(selected_tableau1)
    
    # Sélection du tableau pour le deuxième paramètre
    tableau2_choice = st.selectbox('Choisir un tableau pour le deuxième paramètre', list(tableaux.keys()))
    
    # Récupération du tableau sélectionné pour le deuxième paramètre
    selected_tableau2 = tableaux[tableau2_choice]
    
    # Affichage du tableau pour le deuxième paramètre
    st.write('Tableau sélectionné pour le deuxième paramètre :')
    st.write(selected_tableau2)
    #selectionner les lignes a utiliser
    st.write('Selectionner les lignes a utiliser :')
    #ligne de debut .
    ligne_debut_tableau1 = st.number_input('ligne de debut tableau 1', min_value=1, max_value=1000, value=1)
    #ligne de fin .
    ligne_fin = st.number_input('ligne de fin tableau 1', min_value=1, max_value=1000, value=1)
    taille = ligne_fin - ligne_debut_tableau1
    ligne_debut_tableau2 = st.number_input('ligne de debut tableau 2', min_value=1, max_value=1000, value=1)
    # Récupération des lignes 1 à 12 des deux tableaux pour les paramètres
    param1_values = selected_tableau1.iloc[ligne_debut_tableau1:ligne_fin+1, :]
    param2_values = selected_tableau2.iloc[ligne_debut_tableau2:ligne_debut_tableau2+taille+1, :]
    
    # Sélection des colonnes (paramètres) pour chaque tableau
    param1_columns = param1_values.columns.tolist()
    param2_columns = param2_values.columns.tolist()
    
    # Sélection des paramètres pour le premier paramètre
    selected_param1 = st.selectbox('Choisir un paramètre pour le premier paramètre', param1_columns)
    
    # Sélection des paramètres pour le deuxième paramètre
    selected_param2 = st.selectbox('Choisir un paramètre pour le deuxième paramètre', param2_columns)
    
    # Demande de l'expression mathématique du KPI
    expression = st.text_input("Expression mathématique du KPI (utilisez param1 et param2)", value="param1 / param2")
    
    # Demande de la norme inférieure et supérieure du KPI
    norm_lower = st.number_input("Norme inférieure du KPI", value=0.0)
    norm_upper = st.number_input("Norme supérieure du KPI", value=1.0)

    #choix du nom des kpi
    kpi_name = st.text_input("Nom du KPI", value="KPI")
    # Bouton de calcul
    calculate_button = st.button('Calculer')
    
    # Vérification si le bouton de calcul a été cliqué
    if calculate_button or calculate_clicked:
        calculate_clicked = True
        st.session_state['calculate_clicked'] = calculate_clicked

        # Récupération des valeurs des paramètres sélectionnés
        param1_values = param1_values[selected_param1].tolist()
        param2_values = param2_values[selected_param2].tolist()

        calculEtAffichage(param1_values,param2_values)
        
        # Demande à l'utilisateur s'il souhaite effectuer un nouveau calcul


